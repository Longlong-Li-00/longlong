#!/usr/bin/env python3
"""Check whether a keyword exists in the Excel source and generated JSON files."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data" / "website_content.xlsx"
JSON_DIR = ROOT / "assets" / "data"


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def contains_keyword(value: Any, keyword: str) -> bool:
    text = normalize(value)
    return keyword.casefold() in text.casefold()


def search_excel(keyword: str) -> list[dict[str, Any]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    matches: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [normalize(cell) for cell in next(rows)]
        except StopIteration:
            continue

        for row_index, row in enumerate(rows, start=2):
            row_values = list(row)
            hit_columns = []
            snippets = []
            for col_index, cell in enumerate(row_values):
                if contains_keyword(cell, keyword):
                    header = headers[col_index] if col_index < len(headers) and headers[col_index] else f"col_{col_index + 1}"
                    hit_columns.append(header)
                    snippets.append(f"{header}={normalize(cell)}")

            if hit_columns:
                matches.append(
                    {
                        "sheet": sheet_name,
                        "row": row_index,
                        "columns": hit_columns,
                        "snippet": "; ".join(snippets[:3]),
                    }
                )

    workbook.close()
    return matches


def walk_json(value: Any, keyword: str, path: str = "$") -> list[str]:
    hits: list[str] = []

    if isinstance(value, dict):
        for key, child in value.items():
            hits.extend(walk_json(child, keyword, f"{path}.{key}"))
        return hits

    if isinstance(value, list):
        for index, child in enumerate(value):
            hits.extend(walk_json(child, keyword, f"{path}[{index}]"))
        return hits

    if contains_keyword(value, keyword):
        hits.append(path)
    return hits


def search_json(keyword: str) -> list[dict[str, Any]]:
    matches: list[dict[str, Any]] = []
    for json_path in sorted(JSON_DIR.glob("*.json")):
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        paths = walk_json(payload, keyword)
        if paths:
            matches.append(
                {
                    "file": json_path.name,
                    "count": len(paths),
                    "paths": paths[:5],
                }
            )
    return matches


def main() -> int:
    if len(sys.argv) != 2 or not sys.argv[1].strip():
        print('Usage: python tools\\check_content_update.py "keyword"')
        return 1

    keyword = sys.argv[1].strip()

    if not WORKBOOK_PATH.exists():
        print(f"Excel source not found: {WORKBOOK_PATH}")
        return 1

    if not JSON_DIR.exists():
        print(f"JSON directory not found: {JSON_DIR}")
        return 1

    excel_matches = search_excel(keyword)
    json_matches = search_json(keyword)

    print(f'Keyword: "{keyword}"')
    print()

    print("Excel source:")
    if excel_matches:
        print(f"  Found {len(excel_matches)} matching row(s).")
        for match in excel_matches:
            print(f"  - Sheet: {match['sheet']}, Row: {match['row']}, Columns: {', '.join(match['columns'])}")
            if match["snippet"]:
                print(f"    {match['snippet']}")
    else:
        print("  Not found in Excel.")

    print()
    print("Generated JSON:")
    if json_matches:
        print(f"  Found matches in {len(json_matches)} file(s).")
        for match in json_matches:
            print(f"  - File: {match['file']} ({match['count']} hit(s))")
            for path in match["paths"]:
                print(f"    {path}")
    else:
        print("  Not found in generated JSON.")

    print()
    if excel_matches and not json_matches:
        print("Excel contains the keyword, but generated JSON does not.")
        print("Run: python tools\\generate_site_data.py")
        print("Then check whether the relevant field is skipped or not rendered.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
