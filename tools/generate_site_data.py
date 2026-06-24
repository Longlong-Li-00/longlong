#!/usr/bin/env python3
"""Generate website JSON payloads from data/website_content.xlsx."""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data" / "website_content.xlsx"
OUTPUT_DIR = ROOT / "assets" / "data"


@dataclass(frozen=True)
class SheetSpec:
    required_columns: tuple[str, ...]
    optional_columns: tuple[str, ...] = ()


SHEET_SPECS: dict[str, SheetSpec] = {
    "profile": SheetSpec(("field", "value_en", "value_zh"), ("notes",)),
    "publications": SheetSpec(
        ("id", "year", "title_en", "authors", "journal", "type", "selected"),
        ("title_zh", "volume", "issue", "pages", "doi", "link", "bibtex", "notes"),
    ),
    "patents": SheetSpec(
        ("id", "title_en", "title_zh", "inventors", "applicant", "patent_type"),
        ("date", "patent_number", "application_number", "status", "application_date", "grant_or_publication_date", "notes"),
    ),
    "conferences": SheetSpec(
        ("id", "year", "title_en", "authors", "conference", "presentation_type"),
        ("title_zh", "location", "link", "notes"),
    ),
    "awards": SheetSpec(
        ("id", "year", "name_en", "category"),
        ("name_zh", "level", "organization", "description_en", "description_zh", "notes"),
    ),
    "projects": SheetSpec(
        ("id", "title_en", "title_zh", "visible", "order"),
        (
            "period",
            "lab",
            "background_en",
            "background_zh",
            "methods_en",
            "methods_zh",
            "contribution_en",
            "contribution_zh",
            "result_en",
            "result_zh",
            "image",
            "notes",
        ),
    ),
    "gallery": SheetSpec(
        ("id", "image"),
        ("date", "title_en", "title_zh", "category", "description_en", "description_zh", "source_filename", "notes"),
    ),
    "cv": SheetSpec(
        ("section", "item_id", "order"),
        ("year_or_period", "title_en", "title_zh", "organization", "description_en", "description_zh", "notes"),
    ),
    "links": SheetSpec(("id", "label_en", "label_zh", "url", "type", "visible"), ("notes",)),
}


def log(message: str) -> None:
    print(message)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def clean_text(value: Any) -> str | None:
    if is_blank(value):
        return None
    return str(value).strip()


def clean_bool(value: Any) -> bool | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "yes", "1"}:
            return True
        if normalized in {"false", "no", "0"}:
            return False
    raise ValueError(f"Cannot parse boolean value: {value!r}")


def clean_int(value: Any) -> int | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"Expected integer-like value, got {value!r}")
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    raise ValueError(f"Cannot parse integer value: {value!r}")


def clean_date(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        normalized = value.strip().replace(".", "-").replace("/", "-")
        parts = normalized.split("-")
        if len(parts) == 3 and all(part.isdigit() for part in parts):
            year, month, day = (int(part) for part in parts)
            return date(year, month, day).isoformat()
        return normalized
    raise ValueError(f"Cannot parse date value: {value!r}")


def format_display_date(value: str | None) -> str:
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%Y.%m.%d")
    except ValueError:
        return value


def row_dict(headers: list[str], row_values: tuple[Any, ...]) -> dict[str, Any]:
    return {header: row_values[index] if index < len(row_values) else None for index, header in enumerate(headers)}


def read_sheet_rows(workbook, sheet_name: str, warnings: list[str]) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = [clean_text(cell.value) or "" for cell in sheet[1]]
    spec = SHEET_SPECS[sheet_name]
    missing = [column for column in spec.required_columns if column not in headers]
    if missing:
        raise ValueError(f"Sheet '{sheet_name}' is missing required columns: {', '.join(missing)}")

    expected = set(spec.required_columns) | set(spec.optional_columns)
    extra = [header for header in headers if header and header not in expected]
    if extra:
        warnings.append(f"{sheet_name}: extra columns preserved but not used directly: {', '.join(extra)}")

    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw = row_dict(headers, row)
        if all(is_blank(value) for value in raw.values()):
            continue
        raw["_row_index"] = row_index
        rows.append(raw)
    return rows


def normalize_profile(rows: list[dict[str, Any]]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for row in rows:
        field = clean_text(row["field"])
        if not field:
            raise ValueError(f"profile row {row['_row_index']}: field is required")
        item = {
            "en": clean_text(row.get("value_en")) or "",
            "zh": clean_text(row.get("value_zh")) or "",
            "notes": clean_text(row.get("notes")),
        }
        if field == "research_keywords":
            item["en_list"] = [part.strip() for part in item["en"].split(";") if part.strip()]
            item["zh_list"] = [part.strip() for part in item["zh"].split(";") if part.strip()]
        payload[field] = item
    return payload


def normalize_publications(rows: list[dict[str, Any]], warnings: list[str]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    optional_fields = ("title_zh", "volume", "issue", "pages", "doi", "link", "bibtex")
    for row in rows:
        year = clean_int(row["year"])
        item = {
            "id": clean_text(row["id"]) or "",
            "year": year,
            "title_en": clean_text(row["title_en"]) or "",
            "title_zh": clean_text(row.get("title_zh")),
            "authors": clean_text(row["authors"]) or "",
            "journal": clean_text(row["journal"]) or "",
            "volume": clean_text(row.get("volume")),
            "issue": clean_text(row.get("issue")),
            "pages": clean_text(row.get("pages")),
            "doi": clean_text(row.get("doi")),
            "link": clean_text(row.get("link")),
            "type": (clean_text(row["type"]) or "").lower(),
            "selected": clean_bool(row["selected"]) is True,
            "bibtex": clean_text(row.get("bibtex")),
            "notes": clean_text(row.get("notes")),
        }
        if not item["id"] or year is None or not item["title_en"] or not item["authors"] or not item["journal"] or not item["type"]:
            raise ValueError(f"publications row {row['_row_index']}: missing critical field")
        for field in optional_fields:
            if item[field] in (None, ""):
                warnings.append(f"publications:{item['id']}: optional field '{field}' is blank")
        items.append(item)
    items.sort(key=lambda item: (-item["year"], item["title_en"].lower()))
    return items


def normalize_patents(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        item = {
            "id": clean_text(row["id"]) or "",
            "date": clean_date(row.get("date")),
            "title_en": clean_text(row["title_en"]) or "",
            "title_zh": clean_text(row["title_zh"]) or "",
            "inventors": clean_text(row["inventors"]) or "",
            "applicant": clean_text(row["applicant"]) or "",
            "patent_type": clean_text(row["patent_type"]) or "",
            "patent_number": clean_text(row.get("patent_number")),
            "application_number": clean_text(row.get("application_number")),
            "status": clean_text(row.get("status")),
            "application_date": clean_date(row.get("application_date")),
            "grant_or_publication_date": clean_date(row.get("grant_or_publication_date")),
            "notes": clean_text(row.get("notes")),
        }
        if not item["id"] or not item["title_en"] or not item["patent_type"]:
            raise ValueError(f"patents row {row['_row_index']}: missing critical field")
        if not item["patent_number"] and not item["application_number"]:
            raise ValueError(f"patents row {row['_row_index']}: patent_number or application_number is required")
        item["date_display"] = format_display_date(item["date"])
        item["application_date_display"] = format_display_date(item["application_date"])
        item["grant_or_publication_date_display"] = format_display_date(item["grant_or_publication_date"])
        items.append(item)
    items.sort(
        key=lambda item: (
            item["date"] or "",
            item["application_date"] or "",
            item["title_en"].lower(),
        ),
        reverse=True,
    )
    return items


def normalize_conferences(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        year = clean_int(row["year"])
        item = {
            "id": clean_text(row["id"]) or "",
            "year": year,
            "title_en": clean_text(row["title_en"]) or "",
            "title_zh": clean_text(row.get("title_zh")),
            "authors": (clean_text(row["authors"]) or "").replace("H. Sun", "Haolan Sun"),
            "conference": clean_text(row["conference"]) or "",
            "location": clean_text(row.get("location")),
            "presentation_type": (clean_text(row["presentation_type"]) or "").lower(),
            "link": clean_text(row.get("link")),
            "notes": clean_text(row.get("notes")),
        }
        if not item["id"] or year is None or not item["title_en"] or not item["conference"]:
            raise ValueError(f"conferences row {row['_row_index']}: missing critical field")
        if item["presentation_type"] not in {"oral", "poster"}:
            raise ValueError(f"conferences row {row['_row_index']}: presentation_type must be oral or poster")
        items.append(item)
    items.sort(key=lambda item: (-item["year"], item["title_en"].lower()))
    return items


def normalize_awards(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        year = clean_int(row["year"])
        item = {
            "id": clean_text(row["id"]) or "",
            "year": year,
            "name_en": clean_text(row["name_en"]) or "",
            "name_zh": clean_text(row.get("name_zh")),
            "category": clean_text(row["category"]) or "",
            "level": clean_text(row.get("level")),
            "organization": clean_text(row.get("organization")),
            "description_en": clean_text(row.get("description_en")),
            "description_zh": clean_text(row.get("description_zh")),
            "notes": clean_text(row.get("notes")),
        }
        if not item["id"] or year is None or not item["name_en"] or not item["category"]:
            raise ValueError(f"awards row {row['_row_index']}: missing critical field")
        items.append(item)
    items.sort(key=lambda item: (-item["year"], item["name_en"].lower()))
    return items


def normalize_projects(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        item = {
            "id": clean_text(row["id"]) or "",
            "title_en": clean_text(row["title_en"]) or "",
            "title_zh": clean_text(row["title_zh"]) or "",
            "period": clean_text(row.get("period")),
            "lab": clean_text(row.get("lab")),
            "background_en": clean_text(row.get("background_en")),
            "background_zh": clean_text(row.get("background_zh")),
            "methods_en": clean_text(row.get("methods_en")),
            "methods_zh": clean_text(row.get("methods_zh")),
            "contribution_en": clean_text(row.get("contribution_en")),
            "contribution_zh": clean_text(row.get("contribution_zh")),
            "result_en": clean_text(row.get("result_en")),
            "result_zh": clean_text(row.get("result_zh")),
            "image": clean_text(row.get("image")),
            "visible": clean_bool(row["visible"]) is True,
            "order": clean_int(row["order"]),
            "notes": clean_text(row.get("notes")),
        }
        if not item["id"] or not item["title_en"] or item["order"] is None:
            raise ValueError(f"projects row {row['_row_index']}: missing critical field")
        items.append(item)
    items.sort(key=lambda item: (item["order"], not item["visible"], item["title_en"].lower()))
    return items


def normalize_gallery(rows: list[dict[str, Any]], warnings: list[str]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        iso_date = clean_date(row.get("date"))
        item = {
            "id": clean_text(row["id"]) or "",
            "date": iso_date,
            "date_display": format_display_date(iso_date),
            "title_en": clean_text(row.get("title_en")),
            "title_zh": clean_text(row.get("title_zh")),
            "image": clean_text(row["image"]) or "",
            "category": clean_text(row.get("category")),
            "description_en": clean_text(row.get("description_en")),
            "description_zh": clean_text(row.get("description_zh")),
            "source_filename": clean_text(row.get("source_filename")),
            "notes": clean_text(row.get("notes")),
        }
        if not item["id"] or not item["image"]:
            raise ValueError(f"gallery row {row['_row_index']}: missing critical field")
        if not item["category"]:
            warnings.append(f"gallery:{item['id']}: optional field 'category' is blank")
        items.append(item)
    items.sort(
        key=lambda item: (
            item["date"] or "",
            (item["source_filename"] or item["title_en"] or "").lower(),
        ),
        reverse=True,
    )
    return items


def normalize_cv(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        section = clean_text(row["section"]) or ""
        order = clean_int(row["order"])
        item = {
            "item_id": clean_text(row["item_id"]) or "",
            "year_or_period": clean_text(row.get("year_or_period")),
            "title_en": clean_text(row.get("title_en")),
            "title_zh": clean_text(row.get("title_zh")),
            "organization": clean_text(row.get("organization")),
            "description_en": clean_text(row.get("description_en")),
            "description_zh": clean_text(row.get("description_zh")),
            "order": order,
            "notes": clean_text(row.get("notes")),
        }
        if not section or not item["item_id"] or order is None:
            raise ValueError(f"cv row {row['_row_index']}: missing critical field")
        grouped[section].append(item)
    for section in grouped:
        grouped[section].sort(key=lambda item: (item["order"], item["item_id"]))
    return dict(sorted(grouped.items(), key=lambda pair: pair[0]))


def normalize_links(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        item = {
            "id": clean_text(row["id"]) or "",
            "label_en": clean_text(row["label_en"]) or "",
            "label_zh": clean_text(row["label_zh"]) or "",
            "url": clean_text(row.get("url")),
            "type": clean_text(row["type"]) or "",
            "visible": clean_bool(row["visible"]) is True,
            "notes": clean_text(row.get("notes")),
        }
        if not item["id"] or not item["type"]:
            raise ValueError(f"links row {row['_row_index']}: missing critical field")
        items.append(item)
    items.sort(key=lambda item: (not item["visible"], item["type"], item["id"]))
    return items


def write_json(filename: str, payload: Any) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    warnings: list[str] = []
    log(f"Loading workbook: {WORKBOOK_PATH}")
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    raw_rows = {
        sheet_name: read_sheet_rows(workbook, sheet_name, warnings)
        for sheet_name in SHEET_SPECS
    }

    profile = normalize_profile(raw_rows["profile"])
    publications = normalize_publications(raw_rows["publications"], warnings)
    patents = normalize_patents(raw_rows["patents"])
    conferences = normalize_conferences(raw_rows["conferences"])
    awards = normalize_awards(raw_rows["awards"])
    projects = normalize_projects(raw_rows["projects"])
    gallery = normalize_gallery(raw_rows["gallery"], warnings)
    cv = normalize_cv(raw_rows["cv"])
    links = normalize_links(raw_rows["links"])

    outputs = {
        "profile.json": profile,
        "publications.json": publications,
        "patents.json": patents,
        "conferences.json": conferences,
        "awards.json": awards,
        "projects.json": projects,
        "gallery.json": gallery,
        "cv.json": cv,
        "links.json": links,
    }

    for filename, payload in outputs.items():
        write_json(filename, payload)

    log("Generated JSON files:")
    for filename in outputs:
        log(f"  - assets/data/{filename}")

    log("Record counts:")
    log(f"  - profile fields: {len(profile)}")
    log(f"  - publications: {len(publications)}")
    log(f"  - patents: {len(patents)}")
    log(f"  - conferences: {len(conferences)}")
    log(f"  - awards: {len(awards)}")
    log(f"  - projects: {len(projects)}")
    log(f"  - gallery: {len(gallery)}")
    log(f"  - cv sections: {len(cv)}")
    log(f"  - links: {len(links)}")

    log(f"Validation warnings: {len(warnings)}")
    for warning in warnings:
        log(f"  WARNING: {warning}")

    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
